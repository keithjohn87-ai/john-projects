#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

wb = Workbook()

# Create Instructions sheet first
ins = wb.active
ins.title = 'Instructions'
ins['A1'] = 'Cable Pulling - Master Template'
ins['A1'].font = Font(bold=True)
ins['A2'] = 'SAVE A COPY AS A MASTER TEMPLATE.'
ins['A3'] = 'Save the working file as REEL_<Number>.'
ins['A5'] = 'Input the starting cable footage in A2 of each Phase sheet; A3:A100 will auto-fill (-100 each row).'
ins['A6'] = "Set B2 = 0; B3 will be 100 and increments by 100 down the column. B2 is 0 and B3 is 100.'"
ins['A7'] = 'C2:C100 = tension (ft-lbs) — editable.'
ins['A8'] = 'D2:D100 = actual time (format recommended yyyy-mm-dd hh:mm:ss).'
ins.column_dimensions['A'].width = 100

# Function to create phase sheets
def setup_phase(ws_name):
    ws = wb.create_sheet(title=ws_name)
    # Headers
    ws['A1'] = 'Cable Footage'
    ws['B1'] = 'Intervals'
    ws['C1'] = 'Tension (ft-lbs)'
    ws['D1'] = 'Time'
    for col in ['A','B','C','D']:
        ws.column_dimensions[col].width = 18
    # A2 left blank for starting footage
    ws['A2'] = None
    # B2 = 0
    ws['B2'] = 0
    # Fill formulas down
    for row in range(3, 101):
        a_cell = f'A{row}'
        b_cell = f'B{row}'
        prev_a = f'A{row-1}'
        prev_b = f'B{row-1}'
        ws[a_cell] = f'={prev_a}-100'
        ws[b_cell] = f'={prev_b}+100'
    # Ensure formulas exist for A3 and B3 even if A2 is blank
    # Leave C2:C100 and D2:D100 empty for user input
    # Format D column for datetime
    for row in range(2, 101):
        ws[f'D{row}'].number_format = 'yyyy-mm-dd hh:mm:ss'
    # Unlock C and D input cells
    for row in range(2, 101):
        ws[f'C{row}'].protection = openpyxl.styles.Protection(locked=False)
        ws[f'D{row}'].protection = openpyxl.styles.Protection(locked=False)
    # Protect the sheet to avoid accidental edits to formulas
    ws.protection.sheet = True
    ws.protection.enable = True
    return ws

# Import openpyxl Protection class properly (workaround for references)
import openpyxl
from openpyxl.styles import Protection

# Create Phase sheets
for name in ['Phase A','Phase B','Phase C']:
    ws = setup_phase(name)

# Save workbook
out_path = '/root/.openclaw/workspace/TomLosito.xlsx'
wb.save(out_path)
print('Created', out_path)
