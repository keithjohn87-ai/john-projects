from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

path = '/root/.openclaw/workspace/TomLosito.xlsx'
wb = load_workbook(path)
if 'Instructions' not in wb.sheetnames:
    ins = wb.create_sheet('Instructions', 0)
else:
    ins = wb['Instructions']

ins.delete_rows(1, ins.max_row)

lines = [
    "Cable Pulling - Master Template",
    "",
    "SAVE A COPY AS A MASTER TEMPLATE.",
    "Save the working file as: REEL_<Number> (example: REEL_001.xlsx)",
    "",
    "How to import this data into GoFormz (app.goformz.com):",
    "1) In Excel: File → Save As → choose CSV UTF-8 (one sheet at a time).",
    "   - Note: Save each Phase sheet as a separate CSV if you want separate imports.",
    "2) Log into GoFormz as an Admin and go to Developer / DataSources / Import CSV (or Admin → Import).",
    "3) Upload the CSV file. Map the CSV columns to the target form fields as follows:",
    "   - Cable Footage → Cable Footage",
    "   - Intervals → Intervals",
    "   - Tension (ft-lbs) → Tension",
    "   - Time → Time",
    "4) If you have photos, include public URLs in a PhotoURL column or upload files separately via the GoFormz files area and add file IDs.",
    "5) Run a SMALL test import (5 rows). Verify under Forms → Submissions. If mapping looks correct, run full import.",
    "6) If the import UI asks to match/convert date formats, choose ISO (YYYY-MM-DD) or adjust accordingly.",
    "",
    "Tips:",
    "- Keep the header row intact. Do not include extra header rows or merged cells.",
    "- Use CSV UTF-8 to avoid encoding issues.",
    "- If the foreman needs only 1-3 clicks: they can upload the CSV and accept automatic mapping in most cases.",
    "",
    "If you want, I can also provide a ready-to-send ZIP (CSV + README + one-line email) to forward to the foreman.",
]

for i, line in enumerate(lines, start=1):
    cell = ins.cell(row=i, column=1, value=line)
    cell.alignment = Alignment(wrap_text=True)
    if i == 1:
        cell.font = Font(bold=True, size=14)

ins.column_dimensions['A'].width = 120
wb.save(path)
print('Updated Instructions in', path)
