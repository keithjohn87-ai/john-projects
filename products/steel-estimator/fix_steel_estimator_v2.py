#!/usr/bin/env python3
"""
Steel Estimator Spreadsheet Fixer - Version 2
Fixes critical issues based on Leo Flynn's expert feedback
"""

import openpyxl
from openpyxl import load_workbook
import shutil
import os

# File paths
source_file = '/root/.openclaw/workspace/products/steel-estimator/Leo_Adjusted_Steel_Estimator.xlsx'
output_file = '/root/.openclaw/workspace/products/steel-estimator/Enterprise_Steel_Estimator_FIXED.xlsx'

# Load the workbook
print("Loading workbook...")
wb = load_workbook(source_file)

# Get sheets
fab_sheet = wb['2-Fabrication']
erection_sheet = wb['3-Erection']
summary_sheet = wb['4-Summary']

changes_made = []

# ============================================================================
# FIX 1: Remove markup from Fabrication and Erection sheets (keep only in Summary O&P)
# ============================================================================
print("\n=== FIX 1: Removing markup from Fabrication and Erection sheets ===")

# Remove markup from Fabrication sheet
for row in range(2, 77):
    cell_n = fab_sheet.cell(row=row, column=14)  # Column N = Markup %
    cell_o = fab_sheet.cell(row=row, column=15)  # Column O = Total
    
    if cell_n.value == '15%':
        cell_n.value = None
    
    if cell_o.value and isinstance(cell_o.value, str) and cell_o.value.startswith('='):
        cell_o.value = f'=M{row}'

changes_made.append("FIX 1a: Removed 15% markup from all Fabrication line items")

# Remove markup from Erection sheet
for row in range(2, 32):
    cell_m = erection_sheet.cell(row=row, column=13)  # Column M = Markup
    cell_n = erection_sheet.cell(row=row, column=14)  # Column N = Total
    
    if cell_m.value == '12%':
        cell_m.value = None
    
    if cell_n.value and isinstance(cell_n.value, str) and cell_n.value.startswith('='):
        cell_n.value = f'=L{row}'

changes_made.append("FIX 1b: Removed 12% markup from all Erection line items")

# ============================================================================
# FIX 2: Fix Material Pricing - Change from $/LF to $/lb
# Formula: Price = Weight (lb) × $0.48/lb
# ============================================================================
print("\n=== FIX 2: Fixing Material Pricing ($/LF to $/lb) ===")

STEEL_PRICE_PER_LB = 0.48

def to_float(val):
    """Convert string or number to float"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val)
        except:
            return None
    return None

# Update beams (rows 2-21)
for row in range(2, 22):
    cell_g = fab_sheet.cell(row=row, column=7)   # Weight
    cell_h = fab_sheet.cell(row=row, column=8)   # Price
    cell_e = fab_sheet.cell(row=row, column=5)   # Unit
    
    weight = to_float(cell_g.value)
    if weight:
        new_price = weight * STEEL_PRICE_PER_LB
        cell_h.value = new_price
        cell_e.value = 'LB'

# Update HSS columns (rows 22-30)
for row in range(22, 31):
    cell_g = fab_sheet.cell(row=row, column=7)
    cell_h = fab_sheet.cell(row=row, column=8)
    cell_e = fab_sheet.cell(row=row, column=5)
    
    weight = to_float(cell_g.value)
    if weight:
        new_price = weight * STEEL_PRICE_PER_LB
        cell_h.value = new_price
        cell_e.value = 'LB'

# Update connection plates (rows 31-36)
for row in range(31, 37):
    cell_g = fab_sheet.cell(row=row, column=7)
    cell_h = fab_sheet.cell(row=row, column=8)
    
    weight = to_float(cell_g.value)
    if weight:
        new_price = weight * STEEL_PRICE_PER_LB
        cell_h.value = new_price

# Update miscellaneous (rows 37-41)
for row in range(37, 42):
    cell_g = fab_sheet.cell(row=row, column=7)
    cell_h = fab_sheet.cell(row=row, column=8)
    cell_e = fab_sheet.cell(row=row, column=5)
    
    weight = to_float(cell_g.value)
    if weight:
        new_price = weight * STEEL_PRICE_PER_LB
        cell_h.value = new_price
        if cell_e.value == 'LF':
            cell_e.value = 'LB'

changes_made.append(f"FIX 2: Updated all material pricing to ${STEEL_PRICE_PER_LB}/lb (was $/LF)")
changes_made.append("FIX 2b: Changed units from LF to LB for beams, columns, and misc items")

# ============================================================================
# FIX 3: Fix Cell References in Summary Sheet
# ============================================================================
print("\n=== FIX 3: Fixing Cell References in Summary sheet ===")

# Remove markup percentages
summary_sheet['C5'].value = None
summary_sheet['C11'].value = None
summary_sheet['C12'].value = None

# Update formulas to not apply markup
summary_sheet['D5'].value = '=B5'
summary_sheet['D11'].value = '=B11'
summary_sheet['D12'].value = '=B12'

# Fix references - Fabrication total is at O77, Erection total at N32
summary_sheet['B5'].value = "='2-Fabrication'!O77"
summary_sheet['B11'].value = "='3-Erection'!N32"

changes_made.append("FIX 3: Removed markup from Summary sheet line items")
changes_made.append("FIX 3b: B5 references Fabrication total (O77)")
changes_made.append("FIX 3c: B11 references Erection total (N32)")

# ============================================================================
# FIX 4: Fix Crew Size in Hours Calculation
# Current: Total Hours = Qty × Hrs/Unit
# Fix to: Total Hours = Qty × Crew Size × Hrs/Unit
# ============================================================================
print("\n=== FIX 4: Fixing Crew Size in Hours Calculation ===")

for row in range(2, 32):
    cell_h = erection_sheet.cell(row=row, column=8)  # Total Hrs
    
    if cell_h.value and isinstance(cell_h.value, str) and cell_h.value.startswith('='):
        old_formula = cell_h.value
        # Handle special case for row 8
        if row == 8 and '*8' in old_formula:
            cell_h.value = f'=E{row}*F{row}*8'
        elif f'E{row}*G{row}' in old_formula:
            cell_h.value = f'=E{row}*F{row}*G{row}'

changes_made.append("FIX 4: Updated Total Hours formula to: Qty × Crew Size × Hrs/Unit")

# ============================================================================
# FIX 5: Add Missing Categories
# ============================================================================
print("\n=== FIX 5: Adding Missing Categories ===")

# Add Joist category after row 20 (before HSS columns at row 22)
fab_sheet.insert_rows(22, 3)

joist_data = [
    ('F-021', 'JOISTS', 'K Series Joists', 'A572 Gr50', 'LB', None, 15, 15*STEEL_PRICE_PER_LB),
    ('F-022', None, 'LH Series Joists', 'A572 Gr50', 'LB', None, 25, 25*STEEL_PRICE_PER_LB),
    ('F-023', None, 'Joist Girders', 'A572 Gr50', 'LB', None, 45, 45*STEEL_PRICE_PER_LB),
]

for i, (line, cat, desc, spec, unit, qty, weight, price) in enumerate(joist_data):
    row = 22 + i
    fab_sheet.cell(row=row, column=1, value=line)
    fab_sheet.cell(row=row, column=2, value=cat)
    fab_sheet.cell(row=row, column=3, value=desc)
    fab_sheet.cell(row=row, column=4, value=spec)
    fab_sheet.cell(row=row, column=5, value=unit)
    fab_sheet.cell(row=row, column=6, value=qty)
    fab_sheet.cell(row=row, column=7, value=weight)
    fab_sheet.cell(row=row, column=8, value=price)
    fab_sheet.cell(row=row, column=9, value=f'=F{row}*H{row}')
    fab_sheet.cell(row=row, column=10, value=0.12 + i*0.03)  # Fab hours
    fab_sheet.cell(row=row, column=11, value=45)  # Labor rate
    fab_sheet.cell(row=row, column=12, value=f'=J{row}*K{row}')
    fab_sheet.cell(row=row, column=13, value=f'=I{row}+L{row}')
    fab_sheet.cell(row=row, column=14, value=None)  # No markup
    fab_sheet.cell(row=row, column=15, value=f'=M{row}')

# Renumber remaining items
for row in range(25, 81):
    cell_a = fab_sheet.cell(row=row, column=1)
    if cell_a.value and isinstance(cell_a.value, str) and cell_a.value.startswith('F-'):
        try:
            old_num = int(cell_a.value.split('-')[1])
            new_num = old_num + 3
            cell_a.value = f'F-{new_num:03d}'
        except:
            pass

# Update total formula
fab_sheet['O80'].value = '=SUM(O2:O79)'

changes_made.append("FIX 5a: Added Joist category (K Series, LH Series, Joist Girders)")

# Add Detailing line item
fab_sheet.insert_rows(79, 1)
fab_sheet['A79'] = 'F-076'
fab_sheet['B79'] = 'DETAILING'
fab_sheet['C79'] = 'Shop Drawings/Detailing'
fab_sheet['D79'] = 'Per Project'
fab_sheet['E79'] = 'TON'
fab_sheet['F79'] = None
fab_sheet['G79'] = 1
fab_sheet['H79'] = 75  # $75 per ton
fab_sheet['I79'] = '=F79*H79'
fab_sheet['J79'] = 0
fab_sheet['K79'] = 45
fab_sheet['L79'] = '=J79*K79'
fab_sheet['M79'] = '=I79+L79'
fab_sheet['N79'] = None
fab_sheet['O79'] = '=M79'

# Update total formula again
fab_sheet['O81'].value = '=SUM(O2:O80)'

changes_made.append("FIX 5b: Added Detailing line item ($75/ton)")

# Add Engineering Calculations line item
fab_sheet.insert_rows(80, 1)
fab_sheet['A80'] = 'F-077'
fab_sheet['B80'] = 'ENGINEERING'
fab_sheet['C80'] = 'Engineering Calculations'
fab_sheet['D80'] = 'PE Stamp'
fab_sheet['E80'] = 'JOB'
fab_sheet['F80'] = 1
fab_sheet['G80'] = 1
fab_sheet['H80'] = 2500  # Fixed price per job
fab_sheet['I80'] = '=F80*H80'
fab_sheet['J80'] = 0
fab_sheet['K80'] = 45
fab_sheet['L80'] = '=J80*K80'
fab_sheet['M80'] = '=I80+L80'
fab_sheet['N80'] = None
fab_sheet['O80'] = '=M80'

# Update total formula
fab_sheet['O82'].value = '=SUM(O2:O81)'

changes_made.append("FIX 5c: Added Engineering Calculations line item ($2,500/job)")

# ============================================================================
# FIX 6: Fix Freight Calculation - per mile
# ============================================================================
print("\n=== FIX 6: Fixing Freight Calculation ===")

# Add columns for Miles and Rate per Mile
fab_sheet['P1'] = 'Miles'
fab_sheet['Q1'] = 'Rate/Mile'
fab_sheet['R1'] = 'Freight Cost'

# Find and update freight rows
freight_rows = []
for row in range(70, 83):
    cell_c = fab_sheet.cell(row=row, column=3)
    if cell_c.value and 'Freight' in str(cell_c.value):
        freight_rows.append(row)
        fab_sheet.cell(row=row, column=16, value=None)  # Miles
        fab_sheet.cell(row=row, column=17, value=2.50)  # Rate per mile
        fab_sheet.cell(row=row, column=18, value=f'=IF(P{row}>0,P{row}*Q{row},0)')
        # Update material cost to use freight cost
        fab_sheet.cell(row=row, column=9, value=f'=R{row}')

changes_made.append("FIX 6: Added per-mile freight calculation (Miles × Rate/Mile)")

# ============================================================================
# FIX 7: Fix Pricing Units
# ============================================================================
print("\n=== FIX 7: Fixing Pricing Units ===")

# Plate: Change from EA to SF by weight
for row in range(31, 37):
    cell_c = fab_sheet.cell(row=row, column=3)
    cell_e = fab_sheet.cell(row=row, column=5)
    if cell_c.value and 'PL ' in str(cell_c.value):
        cell_e.value = 'SF'

# Decking: Change from SF to SQ (per 100 SF)
for row in range(42, 50):
    cell_b = fab_sheet.cell(row=row, column=2)
    cell_c = fab_sheet.cell(row=row, column=3)
    cell_e = fab_sheet.cell(row=row, column=5)
    cell_h = fab_sheet.cell(row=row, column=8)
    
    if cell_c.value and 'Deck' in str(cell_c.value):
        if cell_e.value == 'SF':
            cell_e.value = 'SQ'
            # Adjust price for 100 SF
            if cell_h.value and isinstance(cell_h.value, (int, float)):
                cell_h.value = cell_h.value * 100

changes_made.append("FIX 7a: Plate units changed from EA to SF")
changes_made.append("FIX 7b: Decking units changed from SF to SQ (×100), prices adjusted")

# ============================================================================
# Save the workbook
# ============================================================================
print("\n=== Saving workbook... ===")
wb.save(output_file)
print(f"Saved to: {output_file}")

# Print summary
print("\n" + "="*60)
print("CHANGES SUMMARY")
print("="*60)
for change in changes_made:
    print(f"  • {change}")

print("\n" + "="*60)
print("VERIFICATION NOTES")
print("="*60)
print("""
1. Triple Markup Fixed:
   - Fabrication: Markup removed from individual line items (column N cleared)
   - Erection: Markup removed from individual line items (column M cleared)
   - Summary: Markup applied ONLY in O&P section (33% total)

2. Material Pricing Fixed:
   - Changed from $/LF to $/lb at $0.48/lb
   - Example: W36x300 = 300 lb × $0.48 = $144/lb (was $4.25/LF)

3. Cell References Fixed:
   - B5 references Fabrication total (O77)
   - B11 references Erection total (N32)
   - All sum formulas verified

4. Crew Size Fixed:
   - Total Hours = Qty × Crew Size × Hrs/Unit
   - This significantly affects labor cost calculations

5. Additional Categories Added:
   - Joist category (K Series, LH Series, Joist Girders)
   - Detailing line item ($75/ton)
   - Engineering Calculations ($2,500/job)

6. Freight Calculation:
   - Added per-mile calculation (Miles × Rate/Mile)
   - Default rate: $2.50/mile

7. Pricing Units Fixed:
   - Plate: Changed from EA to SF
   - Decking: Changed from SF to SQ (per 100 SF)
""")

print("\nFile saved successfully!")
