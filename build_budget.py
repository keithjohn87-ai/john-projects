from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection

wb = Workbook()
ws = wb.active
ws.title = 'Budget'

# Top title and month input
ws['A1'] = 'Budget - Month'
ws['B1'] = ''  # user can type month here
ws['A1'].font = Font(bold=True, size=14)

# Headers
headers = ['#','Item','Type','Amount','Frequency','Monthly Equivalent','Weekly Equivalent','Due Date','Notes']
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Create 15 bill lines (rows 3..17)
start_row = 3
for i in range(15):
    r = start_row + i
    ws.cell(row=r, column=1, value=i+1)
    ws.cell(row=r, column=2, value='')  # Item
    ws.cell(row=r, column=3, value='Bill')
    ws.cell(row=r, column=4, value='')  # Amount
    ws.cell(row=r, column=5, value='Monthly')  # Frequency default
    # Monthly Equivalent formula in col 6 (F)
    amt = f'D{r}'
    freq = f'E{r}'
    # Formula: IF(E="Monthly",D,IF(E="Weekly",D*52/12,IF(E="Annual",D/12,IF(E="One-time",D/12,D))))
    mformula = f'=IF({freq}="Monthly",{amt},IF({freq}="Weekly",{amt}*52/12,IF({freq}="Annual",{amt}/12,IF({freq}="One-time",{amt}/12,{amt}))))'
    ws.cell(row=r, column=6, value=mformula)
    # Weekly equivalent = Monthly*12/52
    ws.cell(row=r, column=7, value=f'=F{r}*12/52')
    ws.cell(row=r, column=8, value='')
    ws.cell(row=r, column=9, value='')

# Blank row
sep_row = start_row + 15
ws.cell(row=sep_row, column=2, value='')

# Savings lines rows
sav_start = sep_row + 1
for i in range(6):
    r = sav_start + i
    ws.cell(row=r, column=1, value='S'+str(i+1))
    ws.cell(row=r, column=2, value='')
    ws.cell(row=r, column=3, value='Saving')
    ws.cell(row=r, column=4, value='')
    ws.cell(row=r, column=5, value='Monthly')
    amt = f'D{r}'
    freq = f'E{r}'
    mformula = f'=IF({freq}="Monthly",{amt},IF({freq}="Weekly",{amt}*52/12,IF({freq}="Annual",{amt}/12,IF({freq}="One-time",{amt}/12,{amt}))))'
    ws.cell(row=r, column=6, value=mformula)
    ws.cell(row=r, column=7, value=f'=F{r}*12/52')
    ws.cell(row=r, column=8, value='')
    ws.cell(row=r, column=9, value='')

# Summary area to right (columns K-L)
base_col = 11  # K
ws.cell(row=2, column=base_col, value='Summary').font = Font(bold=True)
ws.cell(row=3, column=base_col, value='Monthly Income')
ws.cell(row=3, column=base_col+1, value='')  # income input
ws.cell(row=4, column=base_col, value='Monthly Bills Total')
ws.cell(row=4, column=base_col+1, value=f'=SUM(F{start_row}:F{start_row+14})')
ws.cell(row=5, column=base_col, value='Monthly Savings Total')
ws.cell(row=5, column=base_col+1, value=f'=SUM(F{sav_start}:F{sav_start+5})')
ws.cell(row=6, column=base_col, value='Monthly Total Out')
ws.cell(row=6, column=base_col+1, value=f'={get_column_letter(base_col+1)}4+{get_column_letter(base_col+1)}5')
ws.cell(row=7, column=base_col, value='Leftover (Monthly)')
# Leftover = Income - TotalOut
ws.cell(row=7, column=base_col+1, value=f'={get_column_letter(base_col+1)}3-{get_column_letter(base_col+1)}6')
ws.cell(row=8, column=base_col, value='Leftover (Weekly)')
ws.cell(row=8, column=base_col+1, value=f'={get_column_letter(base_col+1)}7*12/52')
ws.cell(row=9, column=base_col, value='% of Income Used')
ws.cell(row=9, column=base_col+1, value=f'=IF({get_column_letter(base_col+1)}3=0,0,({get_column_letter(base_col+1)}6/{get_column_letter(base_col+1)}3))')

# Format headers widths
col_widths = {'A':6,'B':28,'C':12,'D':12,'E':12,'F':18,'G':18,'H':14,'I':24,'K':18,'L':18}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Unlock input cells: Items (B), Amount (D), Frequency (E), Due (H), Notes (I), Income cell (L3)
# Lock everything then unlock inputs
for row in range(1, sav_start+6):
    for col in range(1,10):
        ws.cell(row=row, column=col).protection = Protection(locked=True)
# Unlock specific ranges
for r in range(start_row, start_row+15):
    ws.cell(row=r, column=2).protection = Protection(locked=False)
    ws.cell(row=r, column=4).protection = Protection(locked=False)
    ws.cell(row=r, column=5).protection = Protection(locked=False)
    ws.cell(row=r, column=8).protection = Protection(locked=False)
    ws.cell(row=r, column=9).protection = Protection(locked=False)
for r in range(sav_start, sav_start+6):
    ws.cell(row=r, column=2).protection = Protection(locked=False)
    ws.cell(row=r, column=4).protection = Protection(locked=False)
    ws.cell(row=r, column=5).protection = Protection(locked=False)
    ws.cell(row=r, column=8).protection = Protection(locked=False)
    ws.cell(row=r, column=9).protection = Protection(locked=False)
# Income input cell
ws.cell(row=3, column=base_col+1).protection = Protection(locked=False)

# Protect sheet
ws.protection.sheet = True
ws.protection.enable = True

# Add a short instruction at top (row 25?) but keep single page — put instruction in top-left second row if desired
ws['A25'] = 'Quick tips: Enter Amount and choose Frequency (Monthly/Weekly/Annual/One-time). Income in Summary is monthly. On mobile, tap cells in Amount or Frequency to edit.'
ws['A25'].alignment = Alignment(wrap_text=True)

out = '/root/.openclaw/workspace/Budget.xlsx'
wb.save(out)
print('Saved', out)
